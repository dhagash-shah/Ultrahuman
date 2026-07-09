#!/usr/bin/env python3
"""
build_data.py — Cashify V4 Growth Huddle data builder.

Reads Growth Huddle Cashify.xlsx and emits / updates data.json.
Parses: Exit Thinking → capital, Public Markets → comps, Con Call Summary → comps,
News Tracker, Special Projects.
The "mis" block is preserved from any prior data.json (written by the cashify-monthly-review skill).

USAGE
    python3 build_data.py growth_huddle.xlsx [data.json]
"""
import os, sys, json, re, datetime
from pathlib import Path
import openpyxl


def norm(v):
    return str(v).strip() if v is not None else ""

def company_name():
    return os.environ.get("COMPANY_NAME", "Ultrahuman")

def find_row(ws, col, predicate, start=1, end=None):
    end = end or ws.max_row
    for r in range(start, end + 1):
        if predicate(norm(ws.cell(r, col).value)):
            return r
    return None

def header_map(ws, row):
    out = {}
    for c in range(1, ws.max_column + 1):
        lab = norm(ws.cell(row, c).value)
        if lab:
            out[lab] = c
    return out

def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, str) and v.strip().upper() in {"NM", "NA", "N/A", "-"}:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None

def parse_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    s = norm(v)
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


# ── Exit Thinking → capital + rounds + cap table ─────────────────────────── #
def parse_exit_thinking(ws):
    out = {}

    hdr = find_row(ws, 2, lambda v: v == "Holding Fund")
    blume = {}
    if hdr:
        hm = header_map(ws, hdr)
        col_cost  = hm.get("Investment Cost")
        col_stake = hm.get("Current Stake")
        col_mtm   = next((c for l, c in hm.items() if "MTM" in l), None)
        col_val   = next((c for l, c in hm.items() if "Blume Valuation" in l), None)
        total_r   = find_row(ws, 2, lambda v: v == "Total", start=hdr + 1)
        if total_r:
            blume = {
                "investmentCost": num(ws.cell(total_r, col_cost).value)   if col_cost  else None,
                "currentStake":   num(ws.cell(total_r, col_stake).value)  if col_stake else None,
                "mtm":            num(ws.cell(total_r, col_mtm).value)    if col_mtm   else None,
                "markedValuation":num(ws.cell(total_r, col_val).value)    if col_val   else None,
            }
        funds = []
        r = hdr + 1
        while r < (total_r or hdr + 1):
            name = norm(ws.cell(r, 2).value)
            if name.lower().startswith("fund"):
                funds.append({
                    "fund":  name,
                    "cost":  num(ws.cell(r, col_cost).value)  if col_cost  else None,
                    "stake": num(ws.cell(r, col_stake).value) if col_stake else None,
                    "mtm":   num(ws.cell(r, col_mtm).value)   if col_mtm   else None,
                })
            r += 1
        blume["funds"] = funds
    out["blume"] = blume

    rnd_hdr = find_row(ws, 2, lambda v: v == "Round")
    rounds, total_primary = [], None
    if rnd_hdr:
        hm = header_map(ws, rnd_hdr)
        c_round = hm.get("Round", 2)
        c_date  = hm.get("Date")
        c_prim  = next((c for l, c in hm.items() if "Primary" in l), None)
        c_pre   = next((c for l, c in hm.items() if "Pre-Money" in l), None)
        c_post  = next((c for l, c in hm.items() if "Post-Money" in l), None)
        r = rnd_hdr + 1
        while r <= ws.max_row:
            label = norm(ws.cell(r, c_round).value)
            if not label:
                r += 1
                if r - rnd_hdr > 40:
                    break
                continue
            if label.lower() == "total":
                total_primary = num(ws.cell(r, c_prim).value) if c_prim else None
                break
            rounds.append({
                "round":     label,
                "date":      parse_date(ws.cell(r, c_date).value) if c_date else None,
                "primary":   num(ws.cell(r, c_prim).value)  if c_prim else None,
                "preMoney":  num(ws.cell(r, c_pre).value)   if c_pre  else None,
                "postMoney": num(ws.cell(r, c_post).value)  if c_post else None,
            })
            r += 1
    if total_primary is None:
        total_primary = round(sum(x["primary"] for x in rounds if x["primary"]), 4) or None
    out["rounds"] = rounds
    out["totalPrimaryRaised"] = total_primary

    dated = [x for x in rounds if x["date"] and x["postMoney"]]
    if dated:
        latest = max(dated, key=lambda x: x["date"])
        out["latestValuation"] = {
            "round": latest["round"], "date": latest["date"],
            "preMoney": latest["preMoney"], "postMoney": latest["postMoney"],
        }
    out["capTable"] = []
    cap_banner = find_row(ws, 2, lambda v: v.lower() == "cap table")
    if cap_banner:
        cap_hdr = find_row(ws, 2, lambda v: v == "Name", start=cap_banner)
        if cap_hdr:
            c_pct = next((c for c in range(3, ws.max_column + 1)
                          if "%" in norm(ws.cell(cap_hdr, c).value)
                          or "Holding" in norm(ws.cell(cap_hdr, c).value)), 3)
            r = cap_hdr + 1
            while r <= ws.max_row:
                name = norm(ws.cell(r, 2).value)
                if not name:
                    break
                pct_v = num(ws.cell(r, c_pct).value)
                out["capTable"].append({"name": name, "pct": pct_v,
                                        "isTotal": name.lower() == "total"})
                if name.lower() == "total":
                    break
                r += 1
    return out


# ── Public Markets → trading comps ───────────────────────────────────────── #
def parse_public_markets(ws):
    def clean_label(v):
        return re.sub(r"\s+", " ", norm(v).lower().replace("\n", " "))

    def find_header():
        candidates = []
        for r in range(1, ws.max_row + 1):
            labels = {clean_label(ws.cell(r, c).value): c for c in range(1, ws.max_column + 1) if norm(ws.cell(r, c).value)}
            if not labels:
                continue
            has_company = "company" in labels or "name" in labels
            has_market_cap = any("market cap" in k or k == "mkt cap" for k in labels)
            has_revenue = any("revenue" in k or "total revenues" in k for k in labels)
            has_valuation = any("m cap / revenue" in k or "mc/rev" in k or "p/s" in k for k in labels)
            if has_company and has_market_cap and has_revenue and has_valuation:
                score = 0
                score += 5 if "company" in labels else 0
                score += 3 if "status" in labels else 0
                score += 2 if "m cap / revenue" in labels else 0
                score += 2 if "m cap / ebitda" in labels else 0
                score += 1 if "gross margin" in labels else 0
                candidates.append((score, r))
        return max(candidates)[1] if candidates else None

    hdr = find_header()
    if not hdr:
        return {}
    hm = header_map(ws, hdr)
    hm_norm = {clean_label(k): c for k, c in hm.items()}

    def col(*aliases):
        wanted = [clean_label(a) for a in aliases]
        for a in wanted:
            if a in hm_norm:
                return hm_norm[a]
        for lab, c in hm_norm.items():
            if any(a in lab for a in wanted):
                return c
        return None

    c_name = col("Company", "Name")
    c_status = col("Status")
    c_market_cap = col("Market Cap", "Mkt Cap")
    c_revenue = col("Revenue", "Total Revenues (LTM)", "Total Revenues")
    c_rev_growth = col("Rev Growth", "Rev Gr.", "Total Revenues/CAGR (1Y TTM)")
    c_gross_margin = col("Gross Margin", "Gross M.", "Gross Profit Margin % (LTM)")
    c_ebitda_margin = col("EBITDA Margin", "EBITDA M.", "EBITDA Margin % (LTM)")
    c_mc_revenue = col("M Cap / Revenue", "MC/Rev", "P/S (LTM)", "PS Ratio")
    c_mc_ebitda = col("M Cap / EBITDA", "MC/EBITDA", "EV/EBITDA (LTM)")

    def row_obj(r):
        get = lambda c: num(ws.cell(r, c).value) if c else None
        return {
            "company":      norm(ws.cell(r, c_name).value),
            "status":       norm(ws.cell(r, c_status).value) if c_status else "",
            "marketCap":    get(c_market_cap),
            "revenue":      get(c_revenue),
            "revGrowth":    get(c_rev_growth),
            "grossMargin":  get(c_gross_margin),
            "ebitdaMargin": get(c_ebitda_margin),
            "mcRevenue":    get(c_mc_revenue),
            "mcEbitda":     get(c_mc_ebitda),
        }

    def score_row(o):
        fields = ("marketCap", "revenue", "revGrowth", "grossMargin",
                  "ebitdaMargin", "mcRevenue", "mcEbitda")
        return sum(1 for k in fields if o.get(k) is not None)

    subject, peers, stats = None, [], {}
    stat_keys = {"peer mean": "mean", "peer median": "median",
                 "peer high": "high", "peer low": "low"}
    r = hdr + 1
    blank_streak = 0
    company_low = company_name().lower()
    while r <= ws.max_row:
        name = norm(ws.cell(r, c_name).value)
        if not name:
            blank_streak += 1
            if blank_streak > 15:
                break
            r += 1
            continue
        blank_streak = 0
        low = name.lower()
        if low in stat_keys:
            o = row_obj(r)
            key = stat_keys[low]
            current = stats.get(key, {})
            candidate = {"mcRevenue": o["mcRevenue"], "mcEbitda": o["mcEbitda"]}
            cand_score = sum(1 for v in candidate.values() if v not in (None, 0))
            current_score = sum(1 for v in current.values() if v not in (None, 0))
            if cand_score >= current_score:
                stats[key] = candidate
        else:
            o = row_obj(r)
            if low == company_low:
                if subject is None or score_row(o) > score_row(subject):
                    subject = o
            else:
                if score_row(o) >= 2:
                    peers.append(o)
        r += 1
    return {"subject": subject, "peers": peers, "stats": stats}


# ── Con Call Summary → last 2 calls per company ───────────────────────────── #
COMPANY_SUFFIX = re.compile(r"\b(ltd\.?|limited|plc|inc\.?|incorporated|group)\b", re.I)

def canon_company(name):
    s = COMPANY_SUFFIX.sub("", name).strip(" .,").lower()
    return re.sub(r"\s+", " ", s)

def parse_con_calls(ws, per_company=2):
    hdr = find_row(ws, 2, lambda v: v == "Company")
    if not hdr:
        return []
    hm = header_map(ws, hdr)
    c_co   = hm.get("Company", 2)
    c_date = hm.get("Date", 3)
    fields = {
        "guidance":      "Guidance",
        "growthDrivers": "Growth Drivers",
        "headwinds":     "Headwinds",
        "kpis":          "KPIs",
    }
    rows = []
    for r in range(hdr + 1, ws.max_row + 1):
        co = norm(ws.cell(r, c_co).value)
        label_a = norm(ws.cell(r, 1).value).lower()
        if not co or label_a in ("prompts", "summary") or co.lower().startswith("whose"):
            continue
        rec = {"company": co, "canon": canon_company(co),
               "date": parse_date(ws.cell(r, c_date).value)}
        for key, label in fields.items():
            col = hm.get(label)
            rec[key] = norm(ws.cell(r, col).value) if col else ""
        rows.append(rec)

    groups = {}
    for rec in rows:
        groups.setdefault(rec["canon"], []).append(rec)

    out = []
    for canon, recs in groups.items():
        dated = [x for x in recs if x["date"]]
        dated.sort(key=lambda x: x["date"], reverse=True)
        seen, deduped = set(), []
        for x in (dated or recs):
            if x["date"] in seen:
                continue
            seen.add(x["date"])
            deduped.append(x)
        chosen = deduped[:per_company]
        display = max((r["company"] for r in recs), key=len)
        out.append({
            "company": display,
            "calls": [{k: v for k, v in c.items() if k != "canon"} for c in chosen],
        })
    out.sort(key=lambda g: (g["calls"][0]["date"] or ""), reverse=True)
    return out


# ── Special Projects ─────────────────────────────────────────────────────── #
def parse_special_projects(ws):
    hdr = find_row(ws, 2, lambda v: v == "Project")
    if not hdr:
        return []
    hm = header_map(ws, hdr)
    c_proj = hm.get("Project", 2)
    c_desc = hm.get("Description")
    c_goal = next((c for l, c in hm.items() if "Goal" in l), None)
    c_done = next((c for l, c in hm.items() if "Completion" in l or "Status" in l), None)
    out = []
    for r in range(hdr + 1, ws.max_row + 1):
        proj = norm(ws.cell(r, c_proj).value)
        if not proj or proj.lower() == "total":
            break
        out.append({
            "project":     proj,
            "description": norm(ws.cell(r, c_desc).value) if c_desc else "",
            "endGoal":     norm(ws.cell(r, c_goal).value) if c_goal else "",
            "status":      norm(ws.cell(r, c_done).value) if c_done else "",
        })
    return out


# ── News Tracker ─────────────────────────────────────────────────────────── #
def parse_news_tracker(ws):
    hdr = find_row(ws, 1, lambda v: v == "Date")
    if not hdr:
        return {"items": [], "industry": [], "companies": []}
    hm = header_map(ws, hdr)
    required = ("URL",)
    if not all(k in hm for k in required):
        return {"items": [], "industry": [], "companies": []}

    def get(r, label):
        col = hm.get(label)
        return norm(ws.cell(r, col).value) if col else ""

    items = []
    for r in range(hdr + 1, ws.max_row + 1):
        summary = get(r, "Summary")
        headline = get(r, "Headline") or summary
        url = get(r, "URL")
        if not headline and not url:
            if r - hdr > 10:
                break
            continue
        show = get(r, "Show on Dashboard")
        if show and show.lower() not in ("yes", "y", "true", "1"):
            continue
        item = {
            "date": parse_date(ws.cell(r, hm.get("Date")).value) if hm.get("Date") else None,
            "type": get(r, "Type"),
            "company": get(r, "Company"),
            "theme": get(r, "Theme"),
            "source": get(r, "Source"),
            "headline": headline,
            "url": url,
            "summary": summary if summary != headline else "",
            "sentiment": get(r, "Sentiment"),
            "relevance": get(r, "Relevance"),
        }
        if item["headline"]:
            items.append(item)

    items.sort(key=lambda x: x["date"] or "", reverse=True)
    industry = [x for x in items if x["type"].lower() == "industry" or not x["company"]]
    companies = [x for x in items if x["type"].lower() == "company" or x["company"]]
    return {"items": items, "industry": industry, "companies": companies}


# ── MIS block — preserved from prior data.json ───────────────────────────── #
def load_mis(out_path):
    if Path(out_path).exists():
        try:
            prev = json.loads(Path(out_path).read_text())
            if prev.get("mis"):
                return prev["mis"]
        except Exception:
            pass
    return None


# ── main ─────────────────────────────────────────────────────────────────── #
def main():
    if len(sys.argv) < 2:
        sys.exit("usage: python3 build_data.py growth_huddle.xlsx [data.json]")
    xlsx     = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else "data.json"

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    def sheet(*names):
        for n in names:
            if n in wb.sheetnames:
                return wb[n]
        return None

    cap_raw = parse_exit_thinking(sheet("Exit Thinking")) if sheet("Exit Thinking") else {}

    FOLDER_MIS_MODEL = "https://drive.google.com/drive/folders/1FPv8GEhseyyfBqo8Miwk918rvMNyOX9y"
    FOLDER_CAP       = "https://drive.google.com/drive/folders/1Fs4N5NXuvB8zxnfsu-Z3FFvD6kytS_ZP"

    data = {
        "meta": {
            "company": company_name(),
            "title": "Growth Huddle",
            "source": Path(xlsx).name,
            "builtFrom": "build_data.py",
        },
        "mis": load_mis(out_path),
        "capital": {
            "blume": cap_raw.get("blume", {}),
            "rounds": cap_raw.get("rounds", []),
            "totalPrimaryRaisedCr": cap_raw.get("totalPrimaryRaised"),
            "latestValuation": cap_raw.get("latestValuation"),
            "capTable": cap_raw.get("capTable", []),
            "links": {
                "roundWise": FOLDER_MIS_MODEL,
                "capTable":  FOLDER_CAP,
                "valuation": FOLDER_CAP,
            },
        },
        "comps": {
            "publicMarkets": parse_public_markets(sheet("Comparables", "Public Markets")) if sheet("Comparables", "Public Markets") else {},
            "conCalls":      parse_con_calls(sheet("Con Call Summary"))    if sheet("Con Call Summary") else [],
        },
        "news": parse_news_tracker(sheet("News Tracker", "Live News Tracker")) if sheet("News Tracker", "Live News Tracker") else {"items": [], "industry": [], "companies": []},
        "specialProjects": parse_special_projects(sheet("Special Projects")) if sheet("Special Projects") else [],
    }

    Path(out_path).write_text(json.dumps(data, indent=2, ensure_ascii=False))
    cap = data["capital"]
    print(f"✓ wrote {out_path}")
    print(f"  mis:             {'present' if data['mis'] else 'MISSING — run cashify-monthly-review skill to populate'}")
    print(f"  rounds:          {len(cap.get('rounds', []))} rounds  ·  total primary ₹{cap.get('totalPrimaryRaisedCr')} Cr")
    print(f"  cap table rows:  {len(cap.get('capTable', []))}")
    print(f"  public peers:    {len(data['comps']['publicMarkets'].get('peers', []))}")
    print(f"  concall cos:     {len(data['comps']['conCalls'])}")
    print(f"  news items:      {len(data['news'].get('items', []))}")
    print(f"  special projects:{len(data['specialProjects'])}")

if __name__ == "__main__":
    main()
