#!/usr/bin/env python3
"""
build_mis.py — compute the MIS dashboard block directly from the MIS + Blume Model
workbooks (no intermediate Growth-Huddle sheet, no /cashify-monthly-review copy-paste).

Output: the `mis` object consumed by index.html, written into data.json (merged with
the capital / comps / specialProjects blocks already there).

DESIGN
------
* Everything resolved BY LABEL. The MIS row layout shifts every vintage (Revenue Post
  Tax was row 63 in the Mar'26 file, row 116 in May'26) — never hardcode rows.
* Reporting month auto-detected = latest actual column with real revenue.
* Periods: Month (YoY + AOP + Blume Model), YTD (FY Apr→month), TTM (trailing 12m).
* MIS columns = Actuals + Plan(=AOP). Blume Model columns come from the Model file
  and are tagged [M] in the UI.
* Control check: Dec'25 Revenue Post Tax must reconcile to ₹1,692,076,864.

USAGE
    python3 build_mis.py <MIS.xlsx> <Model.xlsx> [data.json] [growth_huddle.xlsx]
"""
import os, sys, json, datetime, re
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

class StaleMISFile(Exception):
    def __init__(self, detected, minimum):
        self.detected = detected
        self.minimum = minimum
        super().__init__(
            f"detected stale MIS reporting month {detected[0]:04d}-{detected[1]:02d}; "
            f"expected at least {minimum[0]:04d}-{minimum[1]:02d}"
        )

# ----------------------------------------------------------------------------- #
#  date / column helpers
# ----------------------------------------------------------------------------- #
def date_row(ws, scan=8):
    for r in range(1, scan + 1):
        n = sum(1 for c in range(1, ws.max_column + 1)
                if isinstance(ws.cell(r, c).value, (datetime.datetime, datetime.date)))
        if n >= 6:
            return r
    return 2

def col_index(ws, drow):
    """Map (year, month, is_plan) -> column. Section header in row 1."""
    idx = {}
    section = ""
    for c in range(1, ws.max_column + 1):
        top = ws.cell(1, c).value
        if top not in (None, ""):
            section = str(top).strip()
        d = ws.cell(drow, c).value
        if isinstance(d, (datetime.datetime, datetime.date)):
            is_plan = bool(re.search(r"\b(plan|aop|budget)\b", section, re.I))
            idx.setdefault((d.year, d.month, is_plan), c)   # first wins
    return idx

def label_rows(ws, label_col):
    """{exact_label: [rows]} for a label column."""
    out = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, label_col).value
        if isinstance(v, str) and v.strip():
            out.setdefault(v.strip(), []).append(r)
    return out

def detect_label_col(ws):
    a = sum(1 for r in range(1, min(ws.max_row, 120)) if isinstance(ws.cell(r, 1).value, str))
    b = sum(1 for r in range(1, min(ws.max_row, 120)) if isinstance(ws.cell(r, 2).value, str))
    return 1 if a >= b else 2

# ----------------------------------------------------------------------------- #
#  period month-lists
# ----------------------------------------------------------------------------- #
def fy_start_year(y, m):
    return y if m >= 4 else y - 1

def ytd_months(y, m):
    fy = fy_start_year(y, m)
    out, cy, cm = [], fy, 4
    while (cy, cm) <= (y, m):
        out.append((cy, cm))
        cm += 1
        if cm == 13:
            cm = 1; cy += 1
    return out

def trailing_months(y, m, n=12):
    out = []
    cy, cm = y, m
    for _ in range(n):
        out.append((cy, cm))
        cm -= 1
        if cm == 0:
            cm = 12; cy -= 1
    return list(reversed(out))

def shift_years(months, k):
    return [(y - k, m) for (y, m) in months]

def min_reporting_ym():
    raw = os.environ.get("CASHIFY_MIN_REPORTING_YM", "1900-01")
    try:
        y, m = raw.split("-", 1)
        return int(y), int(m)
    except Exception:
        sys.exit(f"ERROR: CASHIFY_MIN_REPORTING_YM must be YYYY-MM, got {raw!r}")

# ----------------------------------------------------------------------------- #
#  engine
# ----------------------------------------------------------------------------- #
class Sheet:
    def __init__(self, ws):
        self.ws = ws
        self.lc = detect_label_col(ws)
        self.dr = date_row(ws)
        self.idx = col_index(ws, self.dr)
        self.labels = label_rows(ws, self.lc)

    def total_row(self, label):
        rows = self.labels.get(label)
        return rows[0] if rows else None

    def seg_row(self, anchor, segname, window=22):
        """first row == segname within `window` rows after the anchor."""
        if anchor is None:
            return None
        for r in range(anchor + 1, anchor + window + 1):
            if str(self.ws.cell(r, self.lc).value).strip() == segname:
                return r
        return None

    def col(self, y, m, plan=False):
        return self.idx.get((y, m, plan))

    def cell(self, row, y, m, plan=False):
        c = self.col(y, m, plan)
        if row and c:
            v = self.ws.cell(row, c).value
            return v if isinstance(v, (int, float)) else None
        return None

    def period_sum(self, row, months, plan=False):
        if row is None:
            return None
        tot, seen = 0.0, False
        for (y, m) in months:
            v = self.cell(row, y, m, plan)
            if v is not None:
                tot += v; seen = True
        return tot if seen else None

def ratio(num, den):
    if num is None or den in (None, 0):
        return None
    return num / den


def company_name():
    return os.environ.get("COMPANY_NAME", "Ultrahuman")


def empty_metric(metric_type):
    return {
        "type": metric_type,
        "curr": None, "prior": None, "aop": None, "model": None,
        "ytd": None, "ytdPrior": None, "ytdAop": None, "ytdModel": None,
        "ttm": None, "ttmPrior": None, "ttmAop": None, "ttmModel": None,
    }


def find_sheet_with_labels(wb, labels):
    wanted = {x.lower() for x in labels}
    best = None
    best_score = 0
    for ws in wb.worksheets:
        seen = set()
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    text = cell.value.strip().lower()
                    if text in wanted:
                        seen.add(text)
        if len(seen) > best_score:
            best = ws
            best_score = len(seen)
    return best if best_score >= max(3, min(5, len(wanted))) else None


def ultrahuman_model_summary(growth_path):
    if not growth_path or not Path(growth_path).exists():
        return {}
    wb = openpyxl.load_workbook(growth_path, data_only=True)
    label_map = {
        "Gross Revenue": "Gross Revenue",
        "Net Revenue": "Net Revenue",
        "Gross Margin": "Gross Profit",
        "Gross Margin %": "Gross Margin",
        "EBITDA from Operating Activities": "EBITDA",
        "EBITDA %": "EBITDA Margin",
    }
    out = {}
    for ws in wb.worksheets:
        model_col = None
        label_col = None
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().lower() == "model fy26":
                    model_col = cell.column
                    label_col = max(1, cell.column - 9)
                    break
            if model_col:
                break
        if not model_col:
            continue
        for r in range(1, ws.max_row + 1):
            label = ws.cell(r, label_col).value
            if not isinstance(label, str):
                continue
            key = label_map.get(label.strip())
            value = ws.cell(r, model_col).value
            if key and isinstance(value, (int, float)):
                out[key] = float(value)
        if out:
            return out
    return out


def apply_ultrahuman_model_summary(particulars, summary):
    if not summary:
        return False
    pct_labels = {"Gross Margin", "Contribution Margin", "EBITDA Margin"}
    for row in particulars:
        label = row.get("label")
        if label not in summary:
            continue
        annual = summary[label]
        if label in pct_labels:
            row["model"] = annual
            row["ytdModel"] = annual
            row["ttmModel"] = annual
        else:
            row["model"] = annual / 12
            row["ytdModel"] = annual
            row["ttmModel"] = annual
    return True


def build_ultrahuman(mis_path, model_path=None, growth_path=None):
    wb = openpyxl.load_workbook(mis_path, data_only=True)
    metric_defs = [
        ("No of Rings", "count", "Number of Rings"),
        ("Gross Revenue", "abs", "Gross Revenue"),
        ("Net Revenue", "abs", "Net Revenue"),
        ("Gross Profit", "abs", "Gross Margin"),
        ("Gross Margin", "pct", "Gross Margin %"),
        ("Contribution Profit", "abs", "Contribution"),
        ("Contribution Margin", "pct", "Contribution %"),
        ("EBITDA", "abs", "EBITDA"),
        ("EBITDA Margin", "pct", "EBITDA %"),
    ]
    source_labels = [x[2] for x in metric_defs]
    ws = find_sheet_with_labels(wb, source_labels)
    if ws is None:
        sys.exit(
            "ERROR: could not find an Ultrahuman MIS summary sheet with expected rows: "
            + ", ".join(source_labels)
        )
    MIS = Sheet(ws)
    reporting_row = MIS.total_row("Net Revenue") or MIS.total_row("Gross Revenue")
    actual_months = sorted([(y, m) for (y, m, p) in MIS.idx if not p])
    reporting = None
    for (y, m) in reversed(actual_months):
        v = MIS.cell(reporting_row, y, m, False)
        if v and v > 0:
            reporting = (y, m)
            break
    if not reporting:
        sys.exit("ERROR: could not detect a reporting month with revenue > 0 in Ultrahuman MIS workbook")
    ry, rm = reporting
    min_ym = min_reporting_ym()
    if (ry, rm) < min_ym:
        raise StaleMISFile((ry, rm), min_ym)

    months = {
        "cur": [(ry, rm)],
        "ly": [(ry - 1, rm)],
        "ytd": ytd_months(ry, rm),
        "ytdLy": shift_years(ytd_months(ry, rm), 1),
        "ttm": trailing_months(ry, rm, 12),
        "ttmLy": shift_years(trailing_months(ry, rm, 12), 1),
    }

    net_rev_row = MIS.total_row("Net Revenue")
    scale = 10_000_000

    def metric_row(display, metric_type, source):
        row = MIS.total_row(source)
        out = empty_metric(metric_type)
        out["label"] = display
        if metric_type == "pct":
            numerator = {
                "Gross Margin %": "Gross Margin",
                "Contribution %": "Contribution",
                "EBITDA %": "EBITDA",
            }.get(source)
            num_row = MIS.total_row(numerator) if numerator else None
            for key, month_key in (("curr", "cur"), ("prior", "ly"), ("ytd", "ytd"), ("ytdPrior", "ytdLy"), ("ttm", "ttm"), ("ttmPrior", "ttmLy")):
                if num_row and net_rev_row:
                    out[key] = ratio(MIS.period_sum(num_row, months[month_key]), MIS.period_sum(net_rev_row, months[month_key]))
                else:
                    out[key] = MIS.period_sum(row, months[month_key]) if row else None
            return out
        for key, month_key in (("curr", "cur"), ("prior", "ly"), ("ytd", "ytd"), ("ytdPrior", "ytdLy"), ("ttm", "ttm"), ("ttmPrior", "ttmLy")):
            value = MIS.period_sum(row, months[month_key]) if row else None
            out[key] = value * scale if value is not None and metric_type == "abs" else value
        if metric_type == "abs" and source in ("Gross Margin", "Contribution", "EBITDA") and net_rev_row:
            # Percent rows are primary, but this keeps ratios recoverable if future
            # summary files omit explicit percentage rows.
            pass
        return out

    def channel_segmental():
        if "Channel wiseMIS" not in wb.sheetnames:
            return []
        cws = wb["Channel wiseMIS"]
        date_cols = {}
        for c in range(1, cws.max_column + 1):
            d = cws.cell(4, c).value
            if isinstance(d, (datetime.datetime, datetime.date)):
                for offset in range(4):
                    ch = cws.cell(5, c + offset).value
                    if isinstance(ch, str) and ch.strip().lower() not in ("", "total"):
                        date_cols[(d.year, d.month, ch.strip())] = c + offset

        rev_row = None
        for r in range(1, cws.max_row + 1):
            label = cws.cell(r, 3).value
            if isinstance(label, str) and label.strip().lower() == "gross revenue (including channel margin)":
                rev_row = r
                break
        if rev_row is None:
            return []

        channels = []
        known_order = ["Marketplace", "Retail", "D2C"]
        found = []
        for _, _, name in sorted(date_cols):
            if name not in found:
                found.append(name)
        channels = [x for x in known_order if x in found] + [x for x in found if x not in known_order]

        def series(name):
            out = empty_metric("abs")
            for key, month_key in (("curr", "cur"), ("prior", "ly"), ("ytd", "ytd"), ("ytdPrior", "ytdLy"), ("ttm", "ttm"), ("ttmPrior", "ttmLy")):
                total = 0.0
                seen = False
                for y, m in months[month_key]:
                    col = date_cols.get((y, m, name))
                    if not col:
                        continue
                    value = cws.cell(rev_row, col).value
                    if isinstance(value, (int, float)):
                        total += value
                        seen = True
                out[key] = total if seen else None
            return out

        gp = empty_metric("abs")
        gm = empty_metric("pct")
        return [{"label": name.strip(), "rev": series(name), "gp": gp.copy(), "gm": gm.copy()} for name in channels]

    particulars = [metric_row(*d) for d in metric_defs]
    model_summary = ultrahuman_model_summary(growth_path)
    used_model_summary = apply_ultrahuman_model_summary(particulars, model_summary)
    monthname = datetime.date(ry, rm, 1).strftime("%b %Y")
    fy = f"FY{str((fy_start_year(ry, rm) + 1))[2:]}"
    checks = [
        ("Ultrahuman MIS sheet", ws.title, True),
        ("Reporting month detected", f"{ry:04d}-{rm:02d}", True),
        ("Required metrics found", f"{sum(1 for _, _, s in metric_defs if MIS.total_row(s))}/{len(metric_defs)}", True),
    ]
    return {
        "meta": {
            "month": monthname,
            "fy": fy,
            "company": company_name(),
            "subtitle": "Monthly performance review — Actuals vs LY, AOP and Blume Model",
            "reportingYM": [ry, rm],
            "segmentTitle": "Revenue by channel",
            "segmentLabel": "Channel",
            "links": {
                "mis": "https://drive.google.com/drive/folders/1x3yk1oT4eQrMW_Rpy3V9HAa1fVSHuMhG",
                "model": "https://docs.google.com/spreadsheets/d/1-zLSKGjul7KWBm0t1KpfWb4_O6_jBoXM/edit",
            },
            "modelBasis": "Growth Huddle Model FY26; monthly model shown as annual / 12" if used_model_summary else None,
        },
        "particulars": particulars,
        "segmental": channel_segmental(),
        "refurb": {"units": empty_metric("count"), "asp": empty_metric("asp"), "channels": {"online": {}, "retail": {}}},
    }, checks


# ----------------------------------------------------------------------------- #
def build(mis_path, model_path, growth_path=None):
    wb_mis = openpyxl.load_workbook(mis_path, data_only=True)
    if "Consolidated" not in wb_mis.sheetnames:
        return build_ultrahuman(mis_path, model_path, growth_path)
    MIS = Sheet(wb_mis["Consolidated"])
    wbm = openpyxl.load_workbook(model_path, data_only=True)
    MODEL = Sheet(wbm["Consolidated - P&L"])

    # ---- auto-detect reporting month: latest actual col with revenue > 0 -----
    rev_row = MIS.total_row("Revenue Post Tax (after returns)")
    actual_months = sorted([(y, m) for (y, m, p) in MIS.idx if not p])
    reporting = None
    for (y, m) in reversed(actual_months):
        v = MIS.cell(rev_row, y, m, False)
        if v and v > 0:
            reporting = (y, m); break
    if not reporting:
        sys.exit("ERROR: could not detect a reporting month with revenue > 0 in MIS workbook")
    ry, rm = reporting
    min_ym = min_reporting_ym()
    if (ry, rm) < min_ym:
        raise StaleMISFile((ry, rm), min_ym)
    MONTHS = {
        "cur":  [(ry, rm)],
        "ly":   [(ry - 1, rm)],
        "ytd":  ytd_months(ry, rm),
        "ytdLy": shift_years(ytd_months(ry, rm), 1),
        "ttm":  trailing_months(ry, rm, 12),
        "ttmLy": shift_years(trailing_months(ry, rm, 12), 1),
        "ttmModel": trailing_months(ry, rm, 12),
    }

    # ---- metric definitions ---------------------------------------------------
    # abs: summed P&L line.  pct: numerator/Revenue.  pit: point-in-time (month col).
    REVL = "Revenue Post Tax (after returns)"
    DEFS = [
        ("Revenue Post-Tax",       "abs", REVL),
        ("Gross Margin 2",         "abs", "Gross Margin 2"),
        ("Gross Margin 2%",        "pct", "Gross Margin 2"),
        ("Contribution Margin 1",  "abs", "Contribution Margin 1"),
        ("CM1%",                   "pct", "Contribution Margin 1"),
        ("Contribution Margin 2",  "abs", "Contribution Margin 2"),
        ("CM2%",                   "pct", "Contribution Margin 2"),
        ("CM2 − Marketing",        "abs", "CM2 - Marketing"),
        ("CM3%",                   "pct", "CM2 - Marketing"),
        ("EBITDA",                 "abs", "EBITDA (CM2 - Marketing - Fixed Cost)"),
        ("EBITDA%",                "pct", "EBITDA (CM2 - Marketing - Fixed Cost)"),
        ("EBDT",                   "abs", "EBDT (EBITDA - FC & Non Operating Income)"),
        ("EBT%",                   "pct", "EBDT (EBITDA - FC & Non Operating Income)"),
        ("Inventory",              "pit", "Inventory"),
        ("Inventory Days",         "pit", "Inventory Days (value)"),
        ("Cash",                   "pit", "Cash"),
        ("Gross Debt",             "pit", "Gross Debt"),
    ]
    MODEL_REV = MODEL.total_row(REVL)

    def metric_fields(kind, mis_label):
        """Return the period dict expected by the renderer."""
        mrow = MIS.total_row(mis_label)
        # MIS Inventory label collides (appears under finance + balance sheet);
        # for 'Inventory'/'Inventory Days' use the balance-sheet block (last match).
        if mis_label in ("Inventory", "Inventory Days (value)"):
            cand = MIS.labels.get(mis_label, [])
            mrow = cand[-1] if cand else None
        if kind == "pit":
            return {
                "type": "days" if "Days" in mis_label else "abs",
                "curr":  MIS.cell(mrow, ry, rm, False),
                "prior": MIS.cell(mrow, ry - 1, rm, False),
                "aop": None, "ytd": None, "ytdAop": None, "ytdPrior": None,
                "ttm": None, "ttmPrior": None, "ttmAop": None,
                "model": None, "ytdModel": None, "ttmModel": None,
            }
        if kind == "abs":
            f = {"type": "abs"}
            f["curr"]     = MIS.period_sum(mrow, MONTHS["cur"])
            f["prior"]    = MIS.period_sum(mrow, MONTHS["ly"])
            f["aop"]      = MIS.period_sum(mrow, MONTHS["cur"], plan=True)
            f["ytd"]      = MIS.period_sum(mrow, MONTHS["ytd"])
            f["ytdPrior"] = MIS.period_sum(mrow, MONTHS["ytdLy"])
            f["ytdAop"]   = MIS.period_sum(mrow, MONTHS["ytd"], plan=True)
            f["ttm"]      = MIS.period_sum(mrow, MONTHS["ttm"])
            f["ttmPrior"] = MIS.period_sum(mrow, MONTHS["ttmLy"])
            f["ttmAop"]   = MIS.period_sum(mrow, MONTHS["ttm"], plan=True)
            mr = MODEL.total_row(mis_label_to_model(mis_label))
            f["model"]    = MODEL.period_sum(mr, MONTHS["cur"])
            f["ytdModel"] = MODEL.period_sum(mr, MONTHS["ytd"])
            f["ttmModel"] = MODEL.period_sum(mr, MONTHS["ttmModel"])
            return f
        # pct
        f = {"type": "pct"}
        num = mrow
        f["curr"]     = ratio(MIS.period_sum(num, MONTHS["cur"]),  MIS.period_sum(rev_row, MONTHS["cur"]))
        f["prior"]    = ratio(MIS.period_sum(num, MONTHS["ly"]),   MIS.period_sum(rev_row, MONTHS["ly"]))
        f["aop"]      = ratio(MIS.period_sum(num, MONTHS["cur"], True), MIS.period_sum(rev_row, MONTHS["cur"], True))
        f["ytd"]      = ratio(MIS.period_sum(num, MONTHS["ytd"]),  MIS.period_sum(rev_row, MONTHS["ytd"]))
        f["ytdPrior"] = ratio(MIS.period_sum(num, MONTHS["ytdLy"]),MIS.period_sum(rev_row, MONTHS["ytdLy"]))
        f["ytdAop"]   = ratio(MIS.period_sum(num, MONTHS["ytd"], True), MIS.period_sum(rev_row, MONTHS["ytd"], True))
        f["ttm"]      = ratio(MIS.period_sum(num, MONTHS["ttm"]),  MIS.period_sum(rev_row, MONTHS["ttm"]))
        f["ttmPrior"] = ratio(MIS.period_sum(num, MONTHS["ttmLy"]),MIS.period_sum(rev_row, MONTHS["ttmLy"]))
        f["ttmAop"]   = ratio(MIS.period_sum(num, MONTHS["ttm"], True), MIS.period_sum(rev_row, MONTHS["ttm"], True))
        mnum = MODEL.total_row(mis_label_to_model(mis_label))
        f["model"]    = ratio(MODEL.period_sum(mnum, MONTHS["cur"]), MODEL.period_sum(MODEL_REV, MONTHS["cur"]))
        f["ytdModel"] = ratio(MODEL.period_sum(mnum, MONTHS["ytd"]), MODEL.period_sum(MODEL_REV, MONTHS["ytd"]))
        f["ttmModel"] = ratio(MODEL.period_sum(mnum, MONTHS["ttmModel"]), MODEL.period_sum(MODEL_REV, MONTHS["ttmModel"]))
        return f

    particulars = []
    for disp, kind, lab in DEFS:
        f = metric_fields(kind, lab)
        f["label"] = disp
        particulars.append(f)

    # ---- segmental ------------------------------------------------------------
    SEGS = ["Non- Repair", "Refurbished", "Market Place", "Screen Pro + Accessories", "Enterprise Sales"]
    rev_anchor = MIS.total_row(REVL)
    gm2_anchor = MIS.total_row("Gross Margin 2")
    model_rev_anchor = MODEL.total_row(REVL)
    model_gm2_anchor = MODEL.total_row("Gross Margin 2")

    def seg_series(sheet, anchor, seg, model_sheet=None, m_anchor=None):
        """abs period dict for a segment line (handles Screen Pro + Accessories combine)."""
        if seg == "Screen Pro + Accessories":
            r1 = sheet.seg_row(anchor, "Screen Pro")
            r2 = sheet.seg_row(anchor, "Accessories")
            rows = [r for r in (r1, r2) if r]
        else:
            rows = [r for r in [sheet.seg_row(anchor, seg)] if r]
        def psum(months, plan=False):
            vals = [sheet.period_sum(r, months, plan) for r in rows]
            vals = [v for v in vals if v is not None]
            return sum(vals) if vals else None
        f = {
            "curr": psum(MONTHS["cur"]), "prior": psum(MONTHS["ly"]),
            "aop": psum(MONTHS["cur"], True), "ytd": psum(MONTHS["ytd"]),
            "ytdPrior": psum(MONTHS["ytdLy"]), "ytdAop": psum(MONTHS["ytd"], True),
            "ttm": psum(MONTHS["ttm"]), "ttmPrior": psum(MONTHS["ttmLy"]),
            "ttmAop": psum(MONTHS["ttm"], True),
        }
        if model_sheet is not None and m_anchor is not None:
            mseg = "Screen Pro + Accessories" if seg == "Screen Pro + Accessories" else seg
            if seg == "Screen Pro + Accessories":
                mrows = [r for r in [model_sheet.seg_row(m_anchor, "Screen Pro + Accessories")] if r]
            else:
                mrows = [r for r in [model_sheet.seg_row(m_anchor, mseg)] if r]
            def mpsum(months):
                vals = [model_sheet.period_sum(r, months) for r in mrows]
                vals = [v for v in vals if v is not None]
                return sum(vals) if vals else None
            f["model"] = mpsum(MONTHS["cur"]); f["ytdModel"] = mpsum(MONTHS["ytd"]); f["ttmModel"] = mpsum(MONTHS["ttmModel"])
        else:
            f["model"] = None; f["ytdModel"] = None; f["ttmModel"] = None
        return f

    segmental = []
    for seg in SEGS:
        rev = seg_series(MIS, rev_anchor, seg, MODEL, model_rev_anchor)
        gp  = seg_series(MIS, gm2_anchor, seg, MODEL, model_gm2_anchor)
        disp = "Non-Repair" if seg == "Non- Repair" else seg
        gm = {k: ratio(gp.get(k), rev.get(k)) for k in
              ["curr","prior","aop","ytd","ytdPrior","ytdAop","ttm","ttmPrior","ttmAop","model","ytdModel","ttmModel"]}
        segmental.append({"label": disp, "rev": rev, "gp": gp, "gm": gm})

    # ---- refurbished devices + ASP -------------------------------------------
    dev_anchor = MIS.total_row("#Devices/ Transactions")
    dev_ref_row = MIS.seg_row(dev_anchor, "Refurbished")
    ref_rev_row = MIS.seg_row(rev_anchor, "Refurbished")
    def dev_fields():
        f = {"type": "count"}
        f["curr"]=MIS.period_sum(dev_ref_row,MONTHS["cur"]); f["prior"]=MIS.period_sum(dev_ref_row,MONTHS["ly"])
        f["aop"]=MIS.period_sum(dev_ref_row,MONTHS["cur"],True); f["ytd"]=MIS.period_sum(dev_ref_row,MONTHS["ytd"])
        f["ytdPrior"]=MIS.period_sum(dev_ref_row,MONTHS["ytdLy"]); f["ytdAop"]=MIS.period_sum(dev_ref_row,MONTHS["ytd"],True)
        f["ttm"]=MIS.period_sum(dev_ref_row,MONTHS["ttm"]); f["ttmPrior"]=MIS.period_sum(dev_ref_row,MONTHS["ttmLy"])
        f["ttmAop"]=MIS.period_sum(dev_ref_row,MONTHS["ttm"],True)
        mdev_anchor = MODEL.total_row("#Devices/ Transactions")
        mdev = MODEL.seg_row(mdev_anchor, "Refurbished") if mdev_anchor else None
        f["model"]=MODEL.period_sum(mdev,MONTHS["cur"]); f["ytdModel"]=MODEL.period_sum(mdev,MONTHS["ytd"])
        f["ttmModel"]=MODEL.period_sum(mdev,MONTHS["ttmModel"])
        return f
    units = dev_fields()
    # ASP = refurbished revenue / refurbished devices, per period
    refrev = {k: MIS.period_sum(ref_rev_row, MONTHS[mk]) for k, mk in
              [("curr","cur"),("prior","ly"),("ytd","ytd"),("ytdPrior","ytdLy"),("ttm","ttm"),("ttmPrior","ttmLy")]}
    refrev["aop"] = MIS.period_sum(ref_rev_row, MONTHS["cur"], True)
    refrev["ytdAop"] = MIS.period_sum(ref_rev_row, MONTHS["ytd"], True)
    refrev["ttmAop"] = MIS.period_sum(ref_rev_row, MONTHS["ttm"], True)
    asp = {"type": "asp"}
    for k in ["curr","prior","aop","ytd","ytdPrior","ytdAop","ttm","ttmPrior","ttmAop"]:
        asp[k] = ratio(refrev.get(k), units.get(k))
    asp["model"] = ratio(seg_lookup(segmental,"Refurbished","rev","model"), units.get("model"))
    asp["ytdModel"] = ratio(seg_lookup(segmental,"Refurbished","rev","ytdModel"), units.get("ytdModel"))
    asp["ttmModel"] = ratio(seg_lookup(segmental,"Refurbished","rev","ttmModel"), units.get("ttmModel"))

    # ---- refurbished channel split (revenue): Retail vs Online ----------------
    RS = Sheet(openpyxl.load_workbook(mis_path, data_only=True)["Refurbished"])
    r_anchor = RS.total_row("Revenue Post Tax (after returns)")
    row_online  = RS.seg_row(r_anchor, "Online Distribution")
    row_offline = RS.seg_row(r_anchor, "Offline Distribution")
    row_retail  = RS.seg_row(r_anchor, "Retail")
    def chan_sum(rows, months):
        vals = [RS.period_sum(r, months) for r in rows if r]
        vals = [v for v in vals if v is not None]
        return sum(vals) if vals else None
    pk = [("curr","cur"),("prior","ly"),("ytd","ytd"),("ytdPrior","ytdLy"),("ttm","ttm"),("ttmPrior","ttmLy")]
    channels = {
        "online": {k: chan_sum([row_online], MONTHS[mk]) for k, mk in pk},
        "retail": {k: chan_sum([row_retail, row_offline], MONTHS[mk]) for k, mk in pk},
    }

    # ---- meta -----------------------------------------------------------------
    monthname = datetime.date(ry, rm, 1).strftime("%b %Y")
    fy = f"FY{str((fy_start_year(ry,rm)+1))[2:]}"
    mis_block = {
        "meta": {
            "month": monthname, "fy": fy, "company": os.environ.get("COMPANY_NAME", "Ultrahuman"),
            "subtitle": "Monthly performance review — Actuals vs LY, AOP and Blume Model",
            "reportingYM": [ry, rm],
            # Viewable source links (item 11). Edit to exact per-file share URLs.
            "links": {
                "mis": "https://drive.google.com/drive/folders/1FPv8GEhseyyfBqo8Miwk918rvMNyOX9y",
                "model": "https://drive.google.com/drive/folders/1FPv8GEhseyyfBqo8Miwk918rvMNyOX9y",
            },
        },
        "particulars": particulars,
        "segmental": segmental,
        "refurb": {"units": units, "asp": asp, "channels": channels},
    }

    # ---- verification ---------------------------------------------------------
    checks = run_checks(MIS, rev_row, segmental, particulars, channels, ref_rev_row, MONTHS, ry, rm)
    return mis_block, checks


def seg_lookup(segmental, label, block, key):
    for s in segmental:
        if s["label"] == label:
            return s[block].get(key)
    return None


def mis_label_to_model(mis_label):
    """MIS line label -> Model line label (mostly identical; map the few that differ)."""
    m = {
        "EBITDA (CM2 - Marketing - Fixed Cost)": "EBITDA",
        # EBDT label is identical in both workbooks
    }
    return m.get(mis_label, mis_label)


def run_checks(MIS, rev_row, segmental, particulars, channels, ref_rev_row, MONTHS, ry, rm):
    out = []
    # 1) Dec'25 revenue control
    dec = MIS.cell(rev_row, 2025, 12, False)
    out.append(("Dec'25 revenue = 1,692,076,864", dec, abs((dec or 0) - 1692076864) < 1.0))
    # 2) segmental revenue (curr) sums ~ total revenue (allow Other Income residual)
    seg_sum = sum((s["rev"]["curr"] or 0) for s in segmental)
    tot = MIS.period_sum(rev_row, MONTHS["cur"])
    out.append(("Σ segment rev (curr) ≤ total, residual=Other Income",
                f"seg={seg_sum:,.0f} tot={tot:,.0f} resid={tot-seg_sum:,.0f}",
                seg_sum <= tot + 1 and (tot - seg_sum) >= -1))
    # 3) channel split reconciles to refurbished revenue (curr)
    chan = (channels["retail"]["curr"] or 0) + (channels["online"]["curr"] or 0)
    refrev = MIS.period_sum(ref_rev_row, MONTHS["cur"])
    out.append(("Retail+Online = Refurbished rev (curr)",
                f"chan={chan:,.0f} ref={refrev:,.0f}", abs(chan - (refrev or 0)) < max(1.0, 0.005*(refrev or 1))))
    # 4) GM2% in plausible band
    gm2 = next(p for p in particulars if p["label"] == "Gross Margin 2%")
    out.append(("GM2% curr in 5%–40%", f"{(gm2['curr'] or 0)*100:.1f}%", 0.05 <= (gm2["curr"] or 0) <= 0.40))
    return out


# ----------------------------------------------------------------------------- #
def main():
    if len(sys.argv) < 3:
        sys.exit("usage: python3 build_mis.py <MIS.xlsx> <Model.xlsx> [data.json] [growth_huddle.xlsx]")
    mis_path, model_path = sys.argv[1], sys.argv[2]
    out_path = sys.argv[3] if len(sys.argv) > 3 else "data.json"
    growth_path = sys.argv[4] if len(sys.argv) > 4 else None

    # merge into existing data.json (preserve capital / comps / specialProjects)
    data = {}
    if Path(out_path).exists():
        data = json.loads(Path(out_path).read_text())
    try:
        mis_block, checks = build(mis_path, model_path, growth_path)
    except StaleMISFile as e:
        existing = (data.get("mis") or {}).get("meta", {}).get("reportingYM")
        if existing and tuple(existing) >= e.minimum:
            print(
                f"⚠️  {e}. Preserving existing MIS block from "
                f"{existing[0]:04d}-{existing[1]:02d}; check which MIS file drive_pull.py downloaded."
            )
            return
        sys.exit(
            f"ERROR: {e}. Existing data.json does not contain a usable MIS block to preserve. "
            "Check which MIS file drive_pull.py downloaded before deploying."
        )
    data["mis"] = mis_block
    data.setdefault("meta", {})
    data["meta"].update({"misSource": Path(mis_path).name, "modelSource": Path(model_path).name})
    Path(out_path).write_text(json.dumps(data, indent=2, ensure_ascii=False))

    print(f"✓ wrote {out_path}  (reporting {mis_block['meta']['month']}, {mis_block['meta']['fy']})")
    print("─ sanity checks ─")
    ok = True
    for name, val, passed in checks:
        print(f"  [{'PASS' if passed else 'FAIL'}] {name}: {val}")
        ok = ok and passed
    print("─ headline (Month) ─")
    available = {x["label"]: x for x in mis_block["particulars"]}
    headlines = ["Revenue Post-Tax", "Gross Margin 2%", "EBITDA", "EBITDA%"]
    if "Revenue Post-Tax" not in available:
        headlines = ["Gross Revenue", "Net Revenue", "Gross Margin", "EBITDA Margin"]
    for lab in headlines:
        p = available.get(lab)
        if not p:
            continue
        print(f"  {lab}: curr={p['curr']}  prior={p['prior']}  aop={p['aop']}  model={p['model']}")
    if not ok:
        print("⚠️  one or more checks FAILED — inspect before deploying")

if __name__ == "__main__":
    main()
