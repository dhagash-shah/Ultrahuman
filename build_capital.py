#!/usr/bin/env python3
"""
build_capital.py — enrich data.json with the full Capital & Cap Table block.

Sources (all pulled locally via drive_pull.py first):
  captable.pdf            ← shareholding-pattern PDF (diluted %, entity-level)
  investment_profile.xlsx ← Blume fund-wise position (₹ Mn per fund sheet)
  growth_huddle.xlsx      ← Exit Thinking sheet (round-wise, ₹ Cr)
  model.xlsx              ← optional fallback for image-only cap table PDFs

UNITS: everything normalised to ₹ Cr for display (₹ Mn ÷ 10).

USAGE
    python3 build_capital.py captable.pdf investment_profile.xlsx growth_huddle.xlsx [data.json] [model.xlsx]

Run AFTER build_data.py so the rest of data.json already exists.
"""
import os, sys, json, re
from pathlib import Path
import openpyxl

MN_TO_CR = 0.1   # ₹ Mn → ₹ Cr

def company_name():
    return os.environ.get("COMPANY_NAME", "Ultrahuman").strip().lower()

# ── Paste your Google Drive view-links here (shown as buttons in the dashboard) ─
LINKS = {
    "capTable":         "https://drive.google.com/drive/folders/1vn_8fIlOVM5EjtJtG5n1tMmnw2kDZLMz",
    "valuation":        "https://drive.google.com/drive/folders/1vn_8fIlOVM5EjtJtG5n1tMmnw2kDZLMz",
    "roundWise":        "https://drive.google.com/drive/folders/1ILRNcbx54HFP8CNdHkvbdH-hnKFjfPpX",
    "investmentProfile":"https://drive.google.com/drive/folders/1vn_8fIlOVM5EjtJtG5n1tMmnw2kDZLMz",
}
# ────────────────────────────────────────────────────────────────────────────────

# ── Entity name normalisation for cap-table PDF ─────────────────────────── #
NAME_MAP = [
    ("founders",   "Founders",                  "Founders"),
    ("nexus",      "Nexus",                     "Investor"),
    ("awi",        "AWI",                       "Investor"),
    ("deepinder",  "Deepinder Goyal",           "Investor"),
    ("steadview",  "Steadview",                 "Investor"),
    ("flintera",   "Flintera",                  "Investor"),
    ("qualcomm",   "Qualcomm",                  "Investor"),
    ("alteria",    "Alteria",                   "Investor"),
    ("msop",       "MSOP",                      "ESOP"),
    ("mandeep",    "Mandeep Manocha",           "Founders"),
    ("nakul",      "Nakul Kumar",               "Founders"),
    ("amit sethi", "Amit Sethi",                "Founders"),
    ("bessemer",   "Bessemer (BVP)",            "Investor"),
    ("blume",      "Blume Ventures",            "Investor"),
    ("m& s",       "M&S Partners",              "Investor"),
    ("shunwei",    "Shunwei",                   "Investor"),
    ("vivek khare","Vivek Khare",               "Other"),
    ("morningside","Morningside (3C India)",    "Investor"),
    ("3c india",   "Morningside (3C India)",    "Investor"),
    ("aihuishou",  "AiHuiShou (AHS)",           "Investor"),
    ("ahs device", "AiHuiShou (AHS)",           "Investor"),
    ("sin growth", "Sin Growth Partners",       "Investor"),
    ("trifecta",   "Trifecta Venture Debt",     "Investor"),
    ("aep",        "Olympus (AEP)",             "Investor"),
    ("olympus",    "Olympus (AEP)",             "Investor"),
    ("mih",        "MIH / Prosus",              "Investor"),
    ("newquest",   "NewQuest",                  "Investor"),
    ("paramark",   "Paramark / KB",             "Investor"),
    ("amazon",     "Amazon",                    "Investor"),
    ("subodh",     "Subodh Garg",               "Other"),
    ("esop",       "ESOP pool",                 "ESOP"),
]

def clean_name(raw):
    low = raw.lower()
    for kw, disp, grp in NAME_MAP:
        if kw in low:
            return disp, grp
    return re.sub(r"\s+", " ", raw).strip(), "Investor"

def parse_pct(s):
    if not s:
        return None
    m = re.search(r"([\d.]+)\s*%", str(s).replace(" ", ""))
    return float(m.group(1)) / 100 if m else None

def num(v):
    return float(v) if isinstance(v, (int, float)) else None


# ── 1. Cap table from shareholding-pattern PDF ──────────────────────────── #
def parse_captable_pdf(pdf_path):
    try:
        import pdfplumber
    except ImportError:
        sys.exit("Install pdfplumber: pip install pdfplumber")

    with pdfplumber.open(pdf_path) as p:
        rows = []
        for pg in p.pages:
            for tbl in pg.extract_tables():
                rows.extend(tbl)

    hdr_i = next((i for i, r in enumerate(rows)
                  if r and any(str(c).strip() == "S.No" for c in r if c)), None)
    if hdr_i is None:
        raise RuntimeError("cap-table PDF: header row not found")
    hdr   = rows[hdr_i]
    c_sno  = next(i for i, c in enumerate(hdr) if str(c).strip() == "S.No")
    c_name = next(i for i, c in enumerate(hdr) if "Name" in str(c))
    c_pct  = next(i for i, c in enumerate(hdr) if "% holding" in str(c))

    entities, cur = [], None
    for r in rows[hdr_i + 1:]:
        if not r or len(r) <= c_pct:
            continue
        sno = str(r[c_sno]).strip() if r[c_sno] else ""
        nm  = str(r[c_name]).strip() if r[c_name] else ""
        pct = parse_pct(r[c_pct])
        if nm.lower().startswith("total"):
            break
        if sno and nm:
            disp, grp = clean_name(nm)
            cur = {"name": disp, "group": grp, "pct": 0.0, "_raw": nm}
            entities.append(cur)
        if cur and pct:
            cur["pct"] += pct

    merged, order = {}, []
    for e in entities:
        k = e["name"]
        if k not in merged:
            merged[k] = {"name": e["name"], "group": e["group"], "pct": 0.0}
            order.append(k)
        merged[k]["pct"] += e["pct"]
    cap = [merged[k] for k in order]
    for e in cap:
        e["pct"] = round(e["pct"], 4)
    total = round(sum(e["pct"] for e in cap), 4)
    return cap, total


def parse_captable_model(model_path):
    wb = openpyxl.load_workbook(model_path, data_only=True)
    for ws in wb.worksheets:
        stake_row = None
        label_col = None
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().lower() == "stake":
                    stake_row = cell.row
                    label_col = cell.column
                    break
            if stake_row:
                break
        if not stake_row or not label_col:
            continue

        entities = []
        for r in range(max(1, stake_row - 8), stake_row):
            name = ws.cell(r, label_col).value
            if isinstance(name, str) and "blume" in name.lower() and "total" in name.lower():
                raw_vals = [
                    float(ws.cell(r, c).value)
                    for c in range(label_col + 1, min(ws.max_column, label_col + 12) + 1)
                    if isinstance(ws.cell(r, c).value, (int, float)) and 0 <= ws.cell(r, c).value <= 1
                ]
                if raw_vals:
                    pct = raw_vals[3] if len(raw_vals) >= 4 else raw_vals[-1]
                    entities.append({"name": "Blume Ventures", "group": "Investor", "pct": round(pct, 4)})
                break
        for r in range(stake_row + 1, min(ws.max_row, stake_row + 40) + 1):
            name = ws.cell(r, label_col).value
            if not isinstance(name, str) or not name.strip():
                if entities:
                    break
                continue
            raw_vals = []
            for c in range(label_col + 1, min(ws.max_column, label_col + 12) + 1):
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)) and 0 <= v <= 1:
                    raw_vals.append(float(v))
            if not raw_vals:
                continue
            # Prefer the FY26 column when the model carries FY23-FY28 horizontally.
            pct = raw_vals[3] if len(raw_vals) >= 4 else raw_vals[-1]
            disp, grp = clean_name(name)
            entities.append({"name": disp, "group": grp, "pct": round(pct, 4)})

        if entities:
            return entities, round(sum(e["pct"] for e in entities), 4)
    return [], None


# ── 2. Fund-wise holding from Investment Profile workbook (₹ Mn) ─────────── #
def hdr_col(ws, *needles):
    for r in (2, 3, 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                t = re.sub(r"\s+", " ", v).strip().lower()
                if all(n in t for n in needles):
                    return c
    return None

def parse_investment_profile(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    funds = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        crow = None
        for r in range(1, ws.max_row + 1):
            for c in range(1, min(ws.max_column, 10) + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and v.strip().lower() == company_name():
                    crow = r; break
            if crow:
                break
        if not crow:
            continue
        cS  = hdr_col(ws, "stake", "current")
        cP  = hdr_col(ws, "total", "invested", "cost") or hdr_col(ws, "invested", "cost")
        cT  = hdr_col(ws, "total", "value")
        cW  = hdr_col(ws, "moic")
        cK  = hdr_col(ws, "last", "round", "valuation")
        cV  = next((c for r in (2, 3, 1) for c in range(1, ws.max_column + 1)
                    if isinstance(ws.cell(r, c).value, str)
                    and re.sub(r"\s+", " ", ws.cell(r, c).value).strip().lower() == "realised value"), None)
        cX  = hdr_col(ws, "gross", "irr")
        cAF = hdr_col(ws, "status")
        g = lambda c: num(ws.cell(crow, c).value) if c else None
        stake = g(cS)
        funds.append({
            "fund":           sheet,
            "status":         ws.cell(crow, cAF).value if cAF else None,
            "stakeCurrent":   stake,
            "investedCostCr": (g(cP) or 0) * MN_TO_CR if g(cP) is not None else None,
            "mtmCr":          (g(cT) or 0) * MN_TO_CR if g(cT) is not None else None,
            "realisedCr":     (g(cV) or 0) * MN_TO_CR if g(cV) is not None else None,
            "moic":           g(cW),
            "irr":            g(cX),
            "lastRoundValCr": (g(cK) or 0) * MN_TO_CR if g(cK) is not None else None,
        })
    active    = [f for f in funds if isinstance(f["stakeCurrent"], (int, float))]
    agg_stake = sum(f["stakeCurrent"] for f in active)
    agg_cost  = sum((f["investedCostCr"] or 0) for f in active)
    agg_mtm   = sum((f["mtmCr"] or 0) for f in active)
    latest_val = max((f["lastRoundValCr"] for f in funds if f["lastRoundValCr"]), default=None)
    return {
        "currentStake":    agg_stake,
        "investmentCostCr":round(agg_cost, 2),
        "mtmCr":           round(agg_mtm, 2),
        "moic":            round(agg_mtm / agg_cost, 2) if agg_cost else None,
        "latestValuationCr":round(latest_val, 1) if latest_val else None,
        "funds":           funds,
    }


# ── 3. Round-wise details from Growth Huddle (Exit Thinking, ₹ Cr) ─────── #
def parse_rounds_from_gh(gh_xlsx):
    import build_data as bd
    wb = openpyxl.load_workbook(gh_xlsx, data_only=True)
    et = bd.parse_exit_thinking(wb["Exit Thinking"])
    return {
        "rounds":              et.get("rounds", []),
        "totalPrimaryRaisedCr":et.get("totalPrimaryRaised"),
        "latestValuation":     et.get("latestValuation"),
    }


# ── main ─────────────────────────────────────────────────────────────────── #
def main():
    if len(sys.argv) < 4:
        sys.exit("usage: python3 build_capital.py captable.pdf investment_profile.xlsx growth_huddle.xlsx [data.json] [model.xlsx]")
    pdf, ip, gh = sys.argv[1], sys.argv[2], sys.argv[3]
    out_path    = sys.argv[4] if len(sys.argv) > 4 else "data.json"
    model_path  = sys.argv[5] if len(sys.argv) > 5 else None

    try:
        cap, cap_total = parse_captable_pdf(pdf)
    except Exception as e:
        print(f"⚠️  skipped cap-table PDF parse: {e}")
        cap, cap_total = [], None
    if not cap and model_path:
        cap, cap_total = parse_captable_model(model_path)
        if cap:
            print(f"✓ used model workbook cap-table fallback from {Path(model_path).name}")
    blume          = parse_investment_profile(ip)
    rounds_data    = parse_rounds_from_gh(gh)

    capital = {
        "unit":  "₹ Cr",
        "asOf":  "31 Dec 2025",   # update each period
        "blume": blume,
        "rounds":               rounds_data["rounds"],
        "totalPrimaryRaisedCr": rounds_data["totalPrimaryRaisedCr"],
        "latestValuation":      rounds_data["latestValuation"],
        "capTable":             cap,
        "capTableTotalPct":     cap_total,
        "links": {k: v for k, v in LINKS.items() if "PLACEHOLDER" not in v},
    }

    data = json.loads(Path(out_path).read_text()) if Path(out_path).exists() else {}
    data["capital"] = capital
    Path(out_path).write_text(json.dumps(data, indent=2, ensure_ascii=False))

    print(f"✓ wrote capital block to {out_path}")
    total_text = f"{cap_total*100:.2f}%" if cap_total is not None else "not parsed"
    print(f"  cap table: {len(cap)} holders · total {total_text}")
    stake = blume["currentStake"] * 100 if blume.get("currentStake") is not None else None
    print(f"  Blume: stake {stake:.2f}% · cost ₹{blume['investmentCostCr']:.1f} Cr · MTM ₹{blume['mtmCr']:.1f} Cr · {blume['moic']}x" if stake is not None else
          f"  Blume: cost ₹{blume['investmentCostCr']:.1f} Cr · MTM ₹{blume['mtmCr']:.1f} Cr · {blume['moic']}x")
    print(f"  rounds: {len(rounds_data['rounds'])} · total primary ₹{rounds_data['totalPrimaryRaisedCr']} Cr")

if __name__ == "__main__":
    main()
